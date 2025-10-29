# MIT License
#
# Copyright (c) 2024 DALabNOVA
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import pandas as pd
import numpy as np
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers, Color
from openpyxl.utils import get_column_letter
import locale
import colorsys

def extract_median_value(median_str):
    """
    Extract the main median value (not the one in parentheses).
    
    Args:
        median_str: String containing median value, possibly with value in parentheses
        
    Returns:
        float: The extracted median value
    """
    if pd.isna(median_str):
        return None
    
    # Convert to string if not already
    median_str = str(median_str)
    
    # Extract the first number (before any parentheses)
    # Pattern matches numbers with optional decimal point and digits
    match = re.match(r'^\s*([-+]?[0-9]*\.?[0-9]+)', median_str)
    
    if match:
        return float(match.group(1))
    
    return None

def load_excel_data(excel_file="results_test_fitness_size.xlsx", sheet_name=0):
    """
    Load the results from Excel file - both TEST FITNESS and MODEL SIZE tables.
    
    Args:
        excel_file: Path to the Excel file with results
        sheet_name: Sheet name or index to read (0 for first sheet)
        
    Returns:
        tuple: (df_fitness, df_size) - DataFrames for fitness and size tables
    """
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"Results file '{excel_file}' not found.")
    
    try:
        # Read the first sheet with multi-level headers
        df_full = pd.read_excel(excel_file, sheet_name=sheet_name, header=[0, 1])
        print(f"Loaded Excel file: {excel_file}")
        print(f"Sheet: {sheet_name if isinstance(sheet_name, str) else 'First sheet'}")
        print(f"Shape: {df_full.shape}")
        
        # The Excel has multiple tables. We need to find where the first table ends
        # Look for rows where the first column contains "PERFORMANCE - MODEL SIZE"
        first_col = df_full.iloc[:, 0]
        
        # Find the row index where the second table starts
        model_size_idx = None
        for idx, val in enumerate(first_col):
            if pd.notna(val) and 'MODEL SIZE' in str(val).upper():
                model_size_idx = idx
                print(f"Found 'PERFORMANCE - MODEL SIZE' table starting at row {idx}")
                break
        
        # Split into two tables
        if model_size_idx is not None:
            df_fitness = df_full.iloc[:model_size_idx].copy()
            df_size = df_full.iloc[model_size_idx:].copy()
            print(f"TEST FITNESS table shape: {df_fitness.shape}")
            print(f"MODEL SIZE table shape: {df_size.shape}")
        else:
            df_fitness = df_full.copy()
            df_size = None
            print("Warning: MODEL SIZE table not found")
        
        # Process FITNESS table - drop header row
        if len(df_fitness) > 0:
            df_fitness = df_fitness.iloc[1:].reset_index(drop=True)
            print(f"Fitness table after removing header row: {df_fitness.shape}")
        
        # Process SIZE table - drop first 3 rows (separator + header rows)
        if df_size is not None and len(df_size) > 3:
            df_size = df_size.iloc[3:].reset_index(drop=True)
            print(f"Size table after removing header rows: {df_size.shape}")
        
        return df_fitness, df_size
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        raise

def extract_variant_medians(df, variant_name):
    """
    Extract median values for each dataset and model type for a given variant.

    Args:
        df: DataFrame with multi-index columns (variant, statistic)
        variant_name: string like 'VARIANT 20'

    Returns:
        dict: { 'Smallest Model': {dataset_num: median, ...}, ... }
    """
    model_types = ['Smallest Model', 'Optimal Compromise', 'Best Fitness']
    results = {m: {} for m in model_types}

    cols = df.columns
    if not isinstance(cols, pd.MultiIndex):
        raise ValueError("Expected MultiIndex columns in the Excel sheet")

    # Map model types to median column under given variant
    median_col_map = {}
    for model in model_types:
        # Look for columns where:
        # - level0 == variant_name 
        # - level1 contains the model name AND does NOT have .1 suffix (Median column is the first one)
        candidates = [c for c in cols 
                     if str(c[0]).strip() == variant_name 
                     and model in str(c[1]) 
                     and '.1' not in str(c[1]) 
                     and 'Mean' not in str(c[1])]
        
        if candidates:
            median_col_map[model] = candidates[0]
            print(f"  Found Median column for {model}: {candidates[0]}")
        else:
            print(f"  Warning: No Median column found for {model} in {variant_name}")
            median_col_map[model] = None

    # Iterate rows and extract dataset number and medians
    for idx, row in df.iterrows():
        first_cell = row.iloc[0]
        if pd.isna(first_cell):
            continue
        ds_name = str(first_cell)
        
        # Skip rows that contain "PERFORMANCE - TEST FITNESS" or similar headers
        if 'PERFORMANCE' in ds_name.upper() or 'TEST FITNESS' in ds_name.upper():
            continue
        
        m = re.search(r'Dataset\s*(\d+)', ds_name, re.IGNORECASE)
        if not m:
            continue
        ds_num = int(m.group(1))

        for model in model_types:
            col = median_col_map.get(model)
            if col is None:
                continue
            raw = row[col]
            median_val = extract_median_value(raw)
            if median_val is None:
                continue
            results[model][ds_num] = median_val

    return results

def get_dataset_name(df, dataset_num):
    """
    Get the full dataset name from the dataframe.
    
    Args:
        df: DataFrame with the data
        dataset_num: Dataset number (1-15)
        
    Returns:
        str: Dataset name (e.g., 'airfoil', 'bike_sharing')
    """
    for idx, row in df.iterrows():
        first_cell = row.iloc[0]
        if pd.isna(first_cell):
            continue
        ds_name = str(first_cell)
        
        # Skip rows that contain "PERFORMANCE - TEST FITNESS" or similar headers
        if 'PERFORMANCE' in ds_name.upper() or 'TEST FITNESS' in ds_name.upper():
            continue
        
        m = re.search(r'Dataset\s*(\d+)\s*(.+)', ds_name, re.IGNORECASE)
        if m and int(m.group(1)) == dataset_num:
            return m.group(2).strip()
    return ""

def get_all_variants(df):
    """
    Extract all variant names from the DataFrame columns.
    
    Args:
        df: DataFrame with multi-index columns
        
    Returns:
        list: Sorted list of variant names with VARIANT 20 first
    """
    cols = df.columns
    if not isinstance(cols, pd.MultiIndex):
        raise ValueError("Expected MultiIndex columns in the Excel sheet")
    
    # Get unique variant names from level 0 of MultiIndex
    variants = set()
    for col in cols:
        variant_name = str(col[0]).strip()
        # Filter out unwanted column names
        if (variant_name and 
            'Unnamed' not in variant_name and
            'PERFORMANCE' not in variant_name.upper() and
            'TEST FITNESS' not in variant_name.upper()):
            variants.add(variant_name)
    
    # Sort variants: VARIANT 20 first, then others numerically
    def variant_sort_key(v):
        match = re.search(r'VARIANT\s*(\d+)', v, re.IGNORECASE)
        if match:
            num = int(match.group(1))
            # Put VARIANT 20 first (use -1), then sort others numerically
            if num == 20:
                return -1
            return num
        return 999
    
    sorted_variants = sorted(variants, key=variant_sort_key)
    return sorted_variants

def get_color_for_comparison(value, baseline_value, max_improvement_pct=50):
    """
    Get color fill based on comparison with baseline (VARIANT 20).
    
    Args:
        value: Current variant value
        baseline_value: VARIANT 20 baseline value
        max_improvement_pct: Maximum improvement percentage for full color intensity
        
    Returns:
        PatternFill: Color fill object (green if better, red if worse)
    """
    if value is None or baseline_value is None or baseline_value == 0:
        return None
    
    # Calculate percentage difference: negative = improvement (lower error), positive = worse
    diff_pct = ((value - baseline_value) / abs(baseline_value)) * 100
    
    # Normalize to 0-1 scale
    # Cap the intensity at max_improvement_pct in either direction
    intensity = min(abs(diff_pct) / max_improvement_pct, 1.0)
    
    if diff_pct < 0:  # Improvement (lower error) - GREEN
        # Green scale: from white (no improvement) to dark green (max improvement)
        # RGB for green: (0, 255, 0) in full intensity
        # We'll use a lighter green for better visibility
        r = int(144 + (255 - 144) * (1 - intensity))  # 144 to 255
        g = int(238 + (255 - 238) * (1 - intensity))  # 238 to 255
        b = int(144 + (255 - 144) * (1 - intensity))  # 144 to 255
        
        # Convert to hex
        color_hex = f"{r:02X}{g:02X}{b:02X}"
        return PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        
    elif diff_pct > 0:  # Worse (higher error) - RED
        # Red scale: from white (no change) to dark red (max worse)
        r = int(255)
        g = int(255 - int(255 * intensity * 0.7))  # Keep some lightness
        b = int(255 - int(255 * intensity * 0.7))
        
        # Convert to hex
        color_hex = f"{r:02X}{g:02X}{b:02X}"
        return PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    
    return None  # No difference

def create_comparison_excel(df_fitness, df_size, output_file='comparison_results.xlsx'):
    """
    Create an Excel file with all variants - both FITNESS and SIZE tables.
    
    Args:
        df_fitness: DataFrame with the TEST FITNESS data
        df_size: DataFrame with the MODEL SIZE data
        output_file: Output Excel filename
    """
    # Get all variants
    print("\nDetecting all variants in the FITNESS data...")
    all_variants = get_all_variants(df_fitness)
    print(f"Found {len(all_variants)} variants: {', '.join(all_variants)}")
    
    # Extract data for all variants - FITNESS
    print("\n--- Extracting FITNESS data ---")
    variants_fitness_data = {}
    for variant in all_variants:
        print(f"Extracting FITNESS data for {variant}...")
        variants_fitness_data[variant] = extract_variant_medians(df_fitness, variant)
    
    # Extract data for all variants - SIZE
    print("\n--- Extracting SIZE data ---")
    variants_size_data = {}
    if df_size is not None:
        for variant in all_variants:
            print(f"Extracting SIZE data for {variant}...")
            variants_size_data[variant] = extract_variant_medians(df_size, variant)
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    section_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    section_font = Font(bold=True, size=13)
    dataset_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    dataset_font = Font(bold=True, size=11)
    model_font = Font(size=10)
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center", indent=1)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers - column A for dataset names, then one column per variant
    ws['A1'] = ""
    for idx, variant in enumerate(all_variants):
        col_letter = get_column_letter(idx + 2)  # Start from column B
        ws[f'{col_letter}1'] = variant
    
    # Apply header styling to all columns
    for idx in range(len(all_variants) + 1):
        col_letter = get_column_letter(idx + 1)
        cell = ws[f'{col_letter}1']
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Dataset configuration - excluding dataset 4
    dataset_configs = [
        (1, 'airfoil'),
        (2, 'bike_sharing'),
        (3, 'bioavailability'),
        (5, 'breast_cancer'),
        (6, 'concrete_slump'),
        (7, 'concrete_strength'),
        (8, 'diabetes'),
        (9, 'efficiency_cooling'),
        (10, 'efficiency_heating'),
        (11, 'forest_fires'),
        (12, 'istanbul'),
        (13, 'parkinson_updrs'),
        (14, 'ppb'),
        (15, 'resid_build_sale_price')
    ]
    
    model_types_display = {
        'Smallest Model': 'Smallest',
        'Optimal Compromise': 'Best normalized',
        'Best Fitness': 'Best fitness'
    }
    
    current_row = 2
    
    # ========== TABLE 1: TEST FITNESS ==========
    # Write section header
    ws[f'A{current_row}'] = "PERFORMANCE - TEST FITNESS"
    ws[f'A{current_row}'].fill = section_fill
    ws[f'A{current_row}'].font = section_font
    ws[f'A{current_row}'].alignment = left_alignment
    ws[f'A{current_row}'].border = thin_border
    for idx in range(len(all_variants)):
        col_letter = get_column_letter(idx + 2)
        ws[f'{col_letter}{current_row}'].border = thin_border
    current_row += 1
    
    # Iterate through datasets for FITNESS
    for dataset_num, dataset_name in dataset_configs:
        # Get actual dataset name from dataframe if available
        actual_name = get_dataset_name(df_fitness, dataset_num)
        if not actual_name:
            actual_name = dataset_name
        
        # Write dataset header
        dataset_label = f"Dataset {dataset_num} {actual_name}"
        ws[f'A{current_row}'] = dataset_label
        ws[f'A{current_row}'].fill = dataset_fill
        ws[f'A{current_row}'].font = dataset_font
        ws[f'A{current_row}'].alignment = left_alignment
        ws[f'A{current_row}'].border = thin_border
        
        # Apply border to all variant columns for this row
        for idx in range(len(all_variants)):
            col_letter = get_column_letter(idx + 2)
            ws[f'{col_letter}{current_row}'].border = thin_border
        
        current_row += 1
        
        # Write model types and their values (FITNESS)
        for model_type_key, model_type_display in model_types_display.items():
            ws[f'A{current_row}'] = model_type_display
            ws[f'A{current_row}'].font = model_font
            ws[f'A{current_row}'].alignment = left_alignment
            ws[f'A{current_row}'].border = thin_border
            
            # Get baseline value (VARIANT 20, which is the first in all_variants)
            baseline_variant = all_variants[0]  # VARIANT 20 is first
            baseline_val = variants_fitness_data[baseline_variant].get(model_type_key, {}).get(dataset_num)
            
            # Write values for each variant
            for idx, variant in enumerate(all_variants):
                col_letter = get_column_letter(idx + 2)  # Start from column B
                variant_val = variants_fitness_data[variant].get(model_type_key, {}).get(dataset_num)
                
                if variant_val is not None:
                    ws[f'{col_letter}{current_row}'] = round(variant_val, 6)
                    # Force number format with dot as decimal separator
                    ws[f'{col_letter}{current_row}'].number_format = '0.000000'
                    
                    # Apply color only if not the baseline variant (VARIANT 20)
                    if idx > 0 and baseline_val is not None:
                        color_fill = get_color_for_comparison(variant_val, baseline_val)
                        if color_fill:
                            ws[f'{col_letter}{current_row}'].fill = color_fill
                else:
                    ws[f'{col_letter}{current_row}'] = "N/A"
                
                ws[f'{col_letter}{current_row}'].alignment = center_alignment
                ws[f'{col_letter}{current_row}'].border = thin_border
            
            current_row += 1
    
    # Add spacing between tables
    current_row += 2
    
    # ========== TABLE 2: MODEL SIZE ==========
    if df_size is not None and variants_size_data:
        # Write section header
        ws[f'A{current_row}'] = "PERFORMANCE - MODEL SIZE"
        ws[f'A{current_row}'].fill = section_fill
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].alignment = left_alignment
        ws[f'A{current_row}'].border = thin_border
        for idx in range(len(all_variants)):
            col_letter = get_column_letter(idx + 2)
            ws[f'{col_letter}{current_row}'].border = thin_border
        current_row += 1
        
        # Write variant names header row for SIZE table
        ws[f'A{current_row}'] = ""
        for idx, variant in enumerate(all_variants):
            col_letter = get_column_letter(idx + 2)
            ws[f'{col_letter}{current_row}'] = variant
            ws[f'{col_letter}{current_row}'].fill = header_fill
            ws[f'{col_letter}{current_row}'].font = header_font
            ws[f'{col_letter}{current_row}'].alignment = center_alignment
            ws[f'{col_letter}{current_row}'].border = thin_border
        ws[f'A{current_row}'].border = thin_border
        current_row += 1
        
        # Iterate through datasets for SIZE
        for dataset_num, dataset_name in dataset_configs:
            # Get actual dataset name from dataframe if available
            actual_name = get_dataset_name(df_size, dataset_num)
            if not actual_name:
                actual_name = dataset_name
            
            # Write dataset header
            dataset_label = f"Dataset {dataset_num} {actual_name}"
            ws[f'A{current_row}'] = dataset_label
            ws[f'A{current_row}'].fill = dataset_fill
            ws[f'A{current_row}'].font = dataset_font
            ws[f'A{current_row}'].alignment = left_alignment
            ws[f'A{current_row}'].border = thin_border
            
            # Apply border to all variant columns for this row
            for idx in range(len(all_variants)):
                col_letter = get_column_letter(idx + 2)
                ws[f'{col_letter}{current_row}'].border = thin_border
            
            current_row += 1
            
            # Write model types and their values (SIZE) - WITH COLOR (smaller is better)
            for model_type_key, model_type_display in model_types_display.items():
                ws[f'A{current_row}'] = model_type_display
                ws[f'A{current_row}'].font = model_font
                ws[f'A{current_row}'].alignment = left_alignment
                ws[f'A{current_row}'].border = thin_border
                
                # Get baseline value (VARIANT 20, which is the first in all_variants)
                baseline_variant = all_variants[0]  # VARIANT 20 is first
                baseline_val = variants_size_data[baseline_variant].get(model_type_key, {}).get(dataset_num)
                
                # Write values for each variant (SIZE - with color, 1 decimal)
                for idx, variant in enumerate(all_variants):
                    col_letter = get_column_letter(idx + 2)  # Start from column B
                    variant_val = variants_size_data[variant].get(model_type_key, {}).get(dataset_num)
                    
                    if variant_val is not None:
                        ws[f'{col_letter}{current_row}'] = round(variant_val, 1)
                        # Force number format with 1 decimal place
                        ws[f'{col_letter}{current_row}'].number_format = '0.0'
                        
                        # Apply color only if not the baseline variant (VARIANT 20)
                        # Smaller size is better (like fitness)
                        if idx > 0 and baseline_val is not None:
                            color_fill = get_color_for_comparison(variant_val, baseline_val)
                            if color_fill:
                                ws[f'{col_letter}{current_row}'].fill = color_fill
                    else:
                        ws[f'{col_letter}{current_row}'] = "N/A"
                    
                    ws[f'{col_letter}{current_row}'].alignment = center_alignment
                    ws[f'{col_letter}{current_row}'].border = thin_border
                
                current_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    for idx in range(len(all_variants)):
        col_letter = get_column_letter(idx + 2)
        ws.column_dimensions[col_letter].width = 15
    
    # Save the workbook
    wb.save(output_file)
    print(f"\n✓ Excel file created successfully: {output_file}")
    print(f"  Total variants: {len(all_variants)}")
    print(f"  Total datasets: {len(dataset_configs)}")

def main(input_excel="manual_set_results_test_fitness_size.xlsx", 
         output_excel="comparison_results.xlsx"):
    """
    Main function to generate comparison Excel with all variants.
    
    Args:
        input_excel: Path to the input Excel file with results
        output_excel: Path for the output comparison Excel file
    """
    try:
        print("="*80)
        print("EXCEL COMPARISON GENERATOR - ALL VARIANTS")
        print("="*80)
        print(f"\nConfiguration:")
        print(f"  Input file: {input_excel}")
        print(f"  Output file: {output_excel}")
        
        # Load data (both fitness and size tables)
        df_fitness, df_size = load_excel_data(input_excel)
        
        # Create comparison Excel with all variants
        create_comparison_excel(df_fitness, df_size, output_file=output_excel)
        
        print("\n" + "="*80)
        print("PROCESS COMPLETED SUCCESSFULLY")
        print("="*80)
        
    except FileNotFoundError as e:
        print(f"\n✗ Error: {e}")
    except Exception as e:
        print(f"\n✗ An error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    import sys
    
    # Parse command line arguments
    input_excel = "manual_set_results_test_fitness_size.xlsx"
    output_excel = "comparison_results.xlsx"
    
    if len(sys.argv) > 1:
        input_excel = sys.argv[1]
    if len(sys.argv) > 2:
        output_excel = sys.argv[2]
    
    print("\nUsage: python export_comparison_to_excel.py [input_excel] [output_excel]")
    print("Example: python export_comparison_to_excel.py results_test_fitness_size.xlsx comparison.xlsx\n")
    print("Note: All variants found in the input file will be included in the output.\n")
    
    main(input_excel=input_excel, 
         output_excel=output_excel)
